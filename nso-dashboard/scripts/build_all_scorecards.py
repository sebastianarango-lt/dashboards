#!/usr/bin/env python3
"""
build_all_scorecards.py

Generates nso_scorecard_data.json for all NSO studios.
Reads all studio config from nso_config.xlsx (single source of truth).

Per-week fields calculated:
  new_leads, total_leads (cum), presales_week, presales_count (cum),
  cancellations_week, cancellations_count (cum),
  grassroots_leads, grassroots_presales, conversion_rate,
  comm_events, meta_spend, google_spend, grassroots_spend,
  leadteam_fee ($300/active week), total_marketing_spend,
  blended_cpl, blended_cpa, ig_new_followers, est_rmr (null)

Usage:
  python scripts/build_all_scorecards.py
  python scripts/build_all_scorecards.py --dry-run
"""

import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent.parent        # nso-dashboard/
REPO_ROOT = ROOT.parent                    # dashboards/ (where data.json lives)
TODAY = date.today()

CREDS_PATH = ROOT / "credentials" / "service_account.json"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
LEADTEAM_FEE = 300.0

# Studios where the data.json name differs from the display name.
# data.json uses STUDIO_NAME from Snowflake (strip_brand applied).
DATA_NAME_OVERRIDES = {
    "FL-020": "Orlando - Dr Phillips",
}

# Google Sheets tab config per studio code
SHEET_CONFIG = {
    "FL-019": {
        "spend_tab": "*Running Marketing Spend",
        "spend_data_start": 4,      # row 3 = headers, data from row 4
        "spend_date_col": 0,        # DATE OF SPEND
        "spend_amount_col": 2,      # AMOUNT
        "events_tab": "*Events Tracker",
        "events_data_start": 11,    # row 10 = headers
        "events_date_col": 1,       # DATE OF EVENT
    },
    "UT-001": {
        "spend_tab": "Running Promotional Spend",
        "spend_data_start": 3,      # row 2 = headers, data from row 3
        "spend_date_col": 0,
        "spend_amount_col": 2,
        "events_tab": "Presales Event Tracker",
        "events_data_start": 10,    # row 9 = headers
        "events_date_col": 3,       # DATE OF EVENT (4th column: count, name, point person, date)
    },
    "VA-001": {
        # Reston — col layout: NAME | LOCATION | DATE OF EVENT | TYPE | ...
        "spend_tab": "*Running Promotional Spend",
        "spend_data_start": 3,
        "spend_date_col": 0,
        "spend_amount_col": 2,
        "events_tab": "*Events Tracker",
        "events_data_start": 10,
        "events_date_col": 2,  # DATE OF EVENT is col C (index 2)
    },
}

# Naples irregular week date bounds (weeks 1-10 are fixed; 11+ continue Mon-Sun from 4/20)
NAPLES_EXPLICIT_WEEKS = {
    0:  (None,               date(2026, 2,  8)),
    1:  (date(2026, 2,  9),  date(2026, 2, 15)),
    2:  (date(2026, 2, 16),  date(2026, 2, 22)),
    3:  (date(2026, 2, 23),  date(2026, 3,  1)),
    4:  (date(2026, 3,  2),  date(2026, 3,  8)),
    5:  (date(2026, 3,  9),  date(2026, 3, 15)),
    6:  (date(2026, 3, 16),  date(2026, 3, 22)),
    7:  (date(2026, 3, 23),  date(2026, 3, 29)),
    8:  (date(2026, 3, 30),  date(2026, 4,  5)),
    9:  (date(2026, 4,  6),  date(2026, 4, 12)),
    10: (date(2026, 4, 13),  date(2026, 4, 19)),
}
NAPLES_SPECIAL_LABELS = {
    0:  "Pre 2/10",
    1:  "Ads Go Live 2/10 - 2/15",
    24: "Target C/O 7/20 - 7/26",
    27: "Target Grand Open 8/10 - 8/16",
}


# ---------------------------------------------------------------------------
# Date / amount parsing utilities
# ---------------------------------------------------------------------------

def _parse_sheet_date(s, default_year=2026):
    """Parse sheet date strings: 'M/D/YYYY', 'M/D/YY', 'M/D'."""
    s = str(s).strip()
    if not s:
        return None
    parts = s.split("/")
    try:
        if len(parts) == 2:
            return date(default_year, int(parts[0]), int(parts[1]))
        if len(parts) == 3:
            y = int(parts[2])
            if y < 100:
                y += 2000
            return date(y, int(parts[0]), int(parts[1]))
    except (ValueError, TypeError):
        pass
    return None


def _parse_amount(s):
    """Parse '$1,234.56' or '1234.56' → float."""
    cleaned = re.sub(r"[^\d.]", "", str(s))
    try:
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# Config from nso_config.xlsx
# ---------------------------------------------------------------------------

def load_studio_config():
    """Read per-studio config from nso_config.xlsx."""
    wb = openpyxl.load_workbook(ROOT / "nso_config.xlsx")
    ws = wb.active

    studios = []
    for row in range(3, ws.max_row + 1):
        vals = [ws.cell(row, c).value for c in range(1, 23)]
        if not vals[0]:
            continue

        name, code, state = vals[0], vals[1], vals[2]

        def _d(v):
            if isinstance(v, datetime):   # datetime is subclass of date; check first
                return v.date()
            if isinstance(v, date):
                return v
            return None

        week0_date   = _d(vals[3])
        week1_start  = _d(vals[4])
        co_date      = _d(vals[5])
        co_week      = vals[6]
        opening_date = _d(vals[7])
        go_week      = vals[8]
        tier_move    = _d(vals[9])

        total_leads  = float(vals[10]) if vals[10] else None
        presales_tgt = float(vals[11]) if vals[11] else None
        rmr_tgt      = float(vals[12]) if vals[12] else None
        cpl_range    = vals[13]
        cpa_range    = vals[14]
        conv_rate    = float(vals[15]) if vals[15] else None

        fb_page_id   = str(vals[16]) if vals[16] else None
        ig_id        = str(vals[17]) if vals[17] else None
        sheet_url    = vals[18]

        snowflake_id = str(int(vals[19])) if vals[19] else None
        gads_cid     = str(vals[20]) if vals[20] else None
        gbp_loc_id   = str(vals[21]) if vals[21] else None

        studios.append({
            "name":          name,
            "code":          code,
            "state":         state,
            "full_name":     f"SWEAT440 {name}",
            "data_name":     DATA_NAME_OVERRIDES.get(code, name),  # name used in data.json
            "week0_date":    week0_date,
            "week1_start":   week1_start,
            "co_date":       co_date,
            "co_week":       int(co_week) if co_week else None,
            "opening_date":  opening_date,
            "go_week":       int(go_week) if go_week else None,
            "tier_move":     tier_move,
            "targets": {
                "total_leads":       total_leads,
                "presales_count":    presales_tgt,
                "estimated_day1_rmr": rmr_tgt,
                "blended_cpl":       f"${cpl_range}" if cpl_range else None,
                "blended_cpa":       f"${cpa_range}" if cpa_range else None,
                "conversion_rate":   conv_rate,
            },
            "fb_page_id":    fb_page_id,
            "ig_id":         ig_id,
            "sheet_url":     sheet_url,
            "snowflake_id":  snowflake_id,
            "gads_account_id": gads_cid.replace("-", "") if gads_cid else None,
        })

    return studios


# ---------------------------------------------------------------------------
# Week bounds builders
# ---------------------------------------------------------------------------

def make_naples_bounds(num_weeks):
    bounds = []
    explicit = NAPLES_EXPLICIT_WEEKS
    last_end = date(2026, 4, 19)   # Week 10 end
    next_mon = last_end + timedelta(days=1)  # 4/20 = Monday

    bounds.append((0, explicit[0][0], explicit[0][1]))
    for wn in range(1, num_weeks + 1):
        if wn in explicit:
            bounds.append((wn, explicit[wn][0], explicit[wn][1]))
        else:
            offset = wn - 11
            ws = next_mon + timedelta(weeks=offset)
            we = ws + timedelta(days=6)
            bounds.append((wn, ws, we))
    return bounds


def make_mon_sun_bounds(week1_start, num_weeks):
    bounds = [(0, None, week1_start - timedelta(days=1))]
    for i in range(1, num_weeks + 1):
        ws = week1_start + timedelta(weeks=i - 1)
        we = ws + timedelta(days=6)
        bounds.append((i, ws, we))
    return bounds


def fmt_dr(ws, we):
    return f"{ws.month}/{ws.day} - {we.month}/{we.day}"


def current_week_num(bounds):
    cw = 0
    for wn, ws, we in bounds:
        if ws is None:
            continue
        if ws <= TODAY:
            cw = wn
    return cw


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def load_daily_leads(daily_rows):
    """Build {data_name: {date_str: {leads, grassroots_leads}}} from data.json."""
    by = defaultdict(lambda: defaultdict(lambda: {"leads": 0, "grassroots_leads": 0}))
    for r in daily_rows:
        studio = str(r.get("studio", ""))
        d = str(r.get("date", ""))[:10]
        if not studio or len(d) < 10:
            continue
        signups = int(r.get("signups") or 0)
        by[studio][d]["leads"] += signups
        if str(r.get("source", "")).lower() == "grassroots":
            by[studio][d]["grassroots_leads"] += signups
    return by


def load_sales_data(sales_raw):
    """Build {full_name: {date_str: {presales, cancellations, grassroots_presales}}}."""
    by = defaultdict(lambda: defaultdict(
        lambda: {"presales": 0, "cancellations": 0, "grassroots_presales": 0}
    ))
    if not sales_raw:
        return by

    SNOWFLAKE_TO_FULL = {}
    for cfg in _STUDIO_CFG_CACHE:
        if cfg["snowflake_id"]:
            SNOWFLAKE_TO_FULL[cfg["snowflake_id"]] = cfg["full_name"]

    for sid, s in sales_raw.get("studios", {}).items():
        full = SNOWFLAKE_TO_FULL.get(str(sid))
        if not full:
            continue
        for row in s.get("daily", []):
            d = str(row.get("date", ""))[:10]
            src = str(row.get("source", "")).lower()
            by[full][d]["presales"]      += int(row.get("presales") or 0)
            by[full][d]["cancellations"] += int(row.get("cancellations") or 0)
            if src == "grassroots":
                by[full][d]["grassroots_presales"] += int(row.get("presales") or 0)
    return by


def load_ig_followers(social_raw):
    """Build {code: {date_str: follower_count}} from social_insights.json."""
    by = defaultdict(dict)
    for ig in social_raw.get("instagram", []):
        code = ig.get("code", "")
        for r in ig.get("daily", []):
            if "follower_count" in r:
                by[code][r["date"]] = int(r["follower_count"] or 0)
    return by


def load_meta_spend(meta_rows):
    """Build {code: {date_str: spend}} from marketing_data.json meta_ads."""
    all_codes = [cfg["code"] for cfg in _STUDIO_CFG_CACHE]
    by = defaultdict(lambda: defaultdict(float))
    for r in meta_rows:
        cname = str(r.get("campaign_name", ""))
        d = str(r.get("date", ""))[:10]
        spend = float(r.get("spend") or 0)
        if spend <= 0 or len(d) < 10:
            continue
        for code in all_codes:
            if code.lower() in cname.lower():
                by[code][d] += spend
                break
    return by


def load_google_spend(gads_rows):
    """Build {code: {date_str: spend}} from nso_google_ads.json google_ads."""
    by = defaultdict(lambda: defaultdict(float))
    for cfg in _STUDIO_CFG_CACHE:
        acct = cfg.get("gads_account_id")
        if not acct:
            continue
        code = cfg["code"]
        for r in gads_rows:
            if str(r.get("account_id", "")) == acct:
                d = str(r.get("date", ""))[:10]
                spend = float(r.get("spend") or 0)
                if spend > 0 and len(d) >= 10:
                    by[code][d] += spend
    return by


def load_grassroots_data(gc, sheet_url, code):
    """Load spend and events dicts from a studio's Google Sheet."""
    empty = defaultdict(float), defaultdict(int)
    cfg = SHEET_CONFIG.get(code)
    if not cfg or not sheet_url:
        return empty

    try:
        sh = gc.open_by_url(sheet_url)
    except Exception as e:
        print(f"  Warning [{code}]: cannot open sheet ({type(e).__name__}): {e}")
        return empty

    spend_by_date = defaultdict(float)
    try:
        ws = sh.worksheet(cfg["spend_tab"])
        rows = ws.get_all_values()
        dc, ac = cfg["spend_date_col"], cfg["spend_amount_col"]
        for row in rows[cfg["spend_data_start"] - 1:]:
            d = _parse_sheet_date(row[dc] if len(row) > dc else "")
            amt = _parse_amount(row[ac] if len(row) > ac else "")
            if d and amt > 0:
                spend_by_date[d.isoformat()] += amt
        total = sum(spend_by_date.values())
        print(f"  [{code}] Spend sheet: {len(spend_by_date)} dates, ${total:,.2f} total")
    except Exception as e:
        print(f"  Warning [{code}]: spend tab error: {e}")

    events_by_date = defaultdict(int)
    try:
        ws = sh.worksheet(cfg["events_tab"])
        rows = ws.get_all_values()
        dc = cfg["events_date_col"]
        for row in rows[cfg["events_data_start"] - 1:]:
            if not any(str(v).strip() for v in row[:5]):
                continue
            d = _parse_sheet_date(row[dc] if len(row) > dc else "")
            if d:
                events_by_date[d.isoformat()] += 1
        print(f"  [{code}] Events sheet: {sum(events_by_date.values())} events")
    except Exception as e:
        print(f"  Warning [{code}]: events tab error: {e}")

    return spend_by_date, events_by_date


# ---------------------------------------------------------------------------
# Per-week aggregation
# ---------------------------------------------------------------------------

def sum_week_data(wn, ws_date, we_date, data_name, full_name,
                  leads_by_date, sales_by_date, ig_fc,
                  meta_spend_by_date, gads_spend_by_date,
                  gr_spend_by_date, events_by_date):
    """Sum all per-week metrics for a single week bound."""
    leads = gr_leads = 0
    presales = cancellations = gr_presales = 0
    ig_sum = 0; has_ig = False
    meta_spend = gads_spend = gr_spend = comm_events = 0.0

    def _add_day(d_str):
        nonlocal leads, gr_leads, presales, cancellations, gr_presales
        nonlocal ig_sum, has_ig, meta_spend, gads_spend, gr_spend, comm_events

        # Leads from data.json (keyed by short name)
        lv = leads_by_date.get(data_name, {}).get(d_str, {})
        leads    += lv.get("leads", 0)
        gr_leads += lv.get("grassroots_leads", 0)

        # Sales from nso_sales_data.json (keyed by full name)
        sv = sales_by_date.get(full_name, {}).get(d_str, {})
        presales      += sv.get("presales", 0)
        cancellations += sv.get("cancellations", 0)
        gr_presales   += sv.get("grassroots_presales", 0)

        # Instagram followers
        fc = ig_fc.get(d_str)
        if fc is not None:
            ig_sum += fc; has_ig = True

        # Ad spend
        meta_spend  += meta_spend_by_date.get(d_str, 0.0)
        gads_spend  += gads_spend_by_date.get(d_str, 0.0)
        gr_spend    += gr_spend_by_date.get(d_str, 0.0)
        comm_events += events_by_date.get(d_str, 0)

    if ws_date is None:
        # Week 0: all data up to and including we_date
        all_dates = set()
        for lookup in (leads_by_date.get(data_name, {}),
                       sales_by_date.get(full_name, {})):
            all_dates.update(lookup.keys())
        all_dates.update(ig_fc.keys())
        all_dates.update(meta_spend_by_date.keys())
        all_dates.update(gads_spend_by_date.keys())
        all_dates.update(gr_spend_by_date.keys())
        all_dates.update(events_by_date.keys())
        we_str = we_date.isoformat()
        for d_str in sorted(all_dates):
            if d_str <= we_str:
                _add_day(d_str)
    else:
        day = ws_date
        while day <= we_date:
            _add_day(day.isoformat())
            day += timedelta(days=1)

    return {
        "leads":         leads,
        "gr_leads":      gr_leads,
        "presales":      presales,
        "cancellations": cancellations,
        "gr_presales":   gr_presales,
        "ig":            ig_sum if has_ig else None,
        "meta_spend":    round(meta_spend, 2),
        "gads_spend":    round(gads_spend, 2),
        "gr_spend":      round(gr_spend, 2),
        "comm_events":   int(comm_events),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

_STUDIO_CFG_CACHE = []   # populated early so load_sales_data can use it


def main():
    dry_run = "--dry-run" in sys.argv

    # -- Load studio config from Excel
    print("Loading nso_config.xlsx ...")
    studio_cfgs = load_studio_config()
    _STUDIO_CFG_CACHE.extend(studio_cfgs)
    for c in studio_cfgs:
        print(f"  {c['name']} ({c['code']})  GO week={c['go_week']}  "
              f"week1={c['week1_start']}  acct={c['gads_account_id']}")

    # -- Load data files
    print("\nLoading data.json ...")
    data_json_path = REPO_ROOT / "data.json"
    if not data_json_path.exists():
        print(f"  ERROR: {data_json_path} not found")
        sys.exit(1)
    with open(data_json_path, encoding="utf-8") as f:
        raw = json.load(f)
    daily_rows = raw.get("daily_detail", [])
    print(f"  {len(daily_rows)} rows")

    try:
        with open(ROOT / "nso_sales_data.json", encoding="utf-8") as f:
            sales_raw = json.load(f)
        print("Loaded nso_sales_data.json")
    except FileNotFoundError:
        sales_raw = {}
        print("  nso_sales_data.json not found — presales will be 0")

    try:
        with open(ROOT / "social_insights.json", encoding="utf-8") as f:
            social_raw = json.load(f)
        print("Loaded social_insights.json")
    except FileNotFoundError:
        social_raw = {}
        print("  social_insights.json not found — IG data will be null")

    try:
        with open(ROOT / "marketing_data.json", encoding="utf-8") as f:
            mkt_raw = json.load(f)
        meta_rows = mkt_raw.get("meta_ads", [])
        print(f"Loaded marketing_data.json ({len(meta_rows)} meta rows)")
    except FileNotFoundError:
        meta_rows = []
        print("  marketing_data.json not found — meta spend will be 0")

    try:
        with open(ROOT / "nso_google_ads.json", encoding="utf-8") as f:
            gads_raw = json.load(f)
        gads_rows = gads_raw.get("google_ads", [])
        print(f"Loaded nso_google_ads.json ({len(gads_rows)} rows)")
    except FileNotFoundError:
        gads_rows = []
        print("  nso_google_ads.json not found — google spend will be 0")

    # -- Build global lookups
    leads_by_date = load_daily_leads(daily_rows)
    sales_by_date = load_sales_data(sales_raw)
    ig_by_code    = load_ig_followers(social_raw)
    meta_by_code  = load_meta_spend(meta_rows)
    gads_by_code  = load_google_spend(gads_rows)

    # -- Google Sheets (grassroots spend + events)
    gc = None
    if CREDS_PATH.exists():
        try:
            from google.oauth2.service_account import Credentials
            import gspread
            creds = Credentials.from_service_account_file(str(CREDS_PATH), scopes=SCOPES)
            gc = gspread.authorize(creds)
            print("Google Sheets auth OK")
        except Exception as e:
            print(f"  Warning: Google Sheets auth failed: {e}")

    gr_spend_by_code  = {}
    events_by_code    = {}
    for cfg in studio_cfgs:
        code = cfg["code"]
        if gc and cfg.get("sheet_url"):
            print(f"\nLoading Sheets for {cfg['name']} ...")
            sp, ev = load_grassroots_data(gc, cfg["sheet_url"], code)
        else:
            sp, ev = defaultdict(float), defaultdict(int)
        gr_spend_by_code[code] = sp
        events_by_code[code]   = ev

    # -- Build scorecard per studio
    output_studios = []

    for cfg in studio_cfgs:
        full  = cfg["full_name"]
        dname = cfg["data_name"]
        code  = cfg["code"]
        go_wk = cfg["go_week"] or 28
        co_wk = cfg["co_week"]
        num_weeks = max(go_wk + 3, 28)

        print(f"\n{'='*60}\n  {full}  (data_name={dname!r})\n{'='*60}")

        # Build week bounds
        is_naples = cfg["code"] == "FL-019"
        if is_naples:
            bounds = make_naples_bounds(num_weeks)
        else:
            bounds = make_mon_sun_bounds(cfg["week1_start"], num_weeks)

        ig_fc_studio    = ig_by_code.get(code, {})
        meta_spend_d    = meta_by_code.get(code, {})
        gads_spend_d    = gads_by_code.get(code, {})
        gr_spend_d      = gr_spend_by_code.get(code, defaultdict(float))
        events_d        = events_by_code.get(code, defaultdict(int))

        cw_num = current_week_num(bounds)
        print(f"  current_week: {cw_num}  (today={TODAY})")

        cum_leads = cum_presales = cum_canc = 0
        weeks_out = []

        for (wn, ws_date, we_date) in bounds:
            w = sum_week_data(
                wn, ws_date, we_date, dname, full,
                leads_by_date, sales_by_date, ig_fc_studio,
                meta_spend_d, gads_spend_d, gr_spend_d, events_d,
            )

            cum_leads    += w["leads"]
            cum_presales += w["presales"]
            cum_canc     += w["cancellations"]

            # Week label
            if wn == 0:
                wk_label = "Week 0"
                if is_naples:
                    dr = NAPLES_SPECIAL_LABELS.get(0, f"Pre {cfg['week1_start'].month}/{cfg['week1_start'].day}")
                else:
                    d0 = cfg["week1_start"]
                    dr = f"Pre {d0.month}/{d0.day}"
            else:
                wk_label = f"WEEK {wn}"
                if is_naples and wn in NAPLES_SPECIAL_LABELS:
                    dr = NAPLES_SPECIAL_LABELS[wn]
                else:
                    dr = fmt_dr(ws_date, we_date)

            # Cumulative nulls: show null if no data yet (not 0)
            tl = float(cum_leads)    if cum_leads    > 0 else (0.0 if wn == 0 else None)
            pc = float(cum_presales) if cum_presales > 0 else (0.0 if wn == 0 else None)
            cc = float(cum_canc)     if cum_canc     > 0 else (0.0 if wn == 0 else None)

            # Weekly fields: show 0 (or None) only for future weeks
            is_active = ws_date is not None and ws_date <= TODAY

            def _weekly_int(v, fallback=None):
                if v > 0:
                    return v
                return 0 if is_active else fallback

            def _weekly_float(v, fallback=None):
                if v > 0:
                    return round(v, 2)
                return 0.0 if is_active else fallback

            # Conversion rate: presales / leads (%)
            conv_rate = None
            if w["leads"] > 0 and is_active:
                conv_rate = round(w["presales"] / w["leads"] * 100, 1)

            # LeadTeam fee: $300 for every active week (wn >= 1)
            leadteam_fee = LEADTEAM_FEE if (is_active and wn >= 1) else None

            # Total marketing spend
            meta_sp  = _weekly_float(w["meta_spend"])
            gads_sp  = _weekly_float(w["gads_spend"])
            gr_sp    = _weekly_float(w["gr_spend"])

            if is_active and wn >= 1:
                total_spend = round(
                    (meta_sp or 0) + (gads_sp or 0) + (gr_sp or 0) + LEADTEAM_FEE, 2
                )
            else:
                total_spend = None

            # CPL / CPA
            blended_cpl = blended_cpa = None
            if total_spend is not None and w["leads"] > 0:
                blended_cpl = round(total_spend / w["leads"], 2)
            if total_spend is not None and w["presales"] > 0:
                blended_cpa = round(total_spend / w["presales"], 2)

            entry = {
                "week":               wk_label,
                "date_range":         dr,
                "date_start":         ws_date.isoformat() if ws_date else None,
                "date_end":           we_date.isoformat(),
                "new_leads":          _weekly_int(w["leads"]),
                "total_leads":        tl,
                "presales_week":      _weekly_int(w["presales"]),
                "presales_count":     pc,
                "cancellations_week": _weekly_int(w["cancellations"]),
                "cancellations_count": cc,
                "grassroots_leads":   _weekly_int(w["gr_leads"]),
                "grassroots_presales": _weekly_int(w["gr_presales"]),
                "conversion_rate":    conv_rate,
                "comm_events":        _weekly_int(w["comm_events"]),
                "ig_new_followers":   w["ig"],
                "meta_spend":         meta_sp if (is_active and wn >= 1) else None,
                "google_spend":       gads_sp if (is_active and wn >= 1) else None,
                "grassroots_spend":   gr_sp if (is_active and wn >= 1) else None,
                "leadteam_fee":       leadteam_fee,
                "total_marketing_spend": total_spend,
                "blended_cpl":        blended_cpl,
                "blended_cpa":        blended_cpa,
                "estimated_day1_rmr": None,
            }

            ig_str = str(w["ig"]) if w["ig"] is not None else "-"
            print(
                f"  W{wn:02d} {dr[:22]:<23} "
                f"+{w['leads']:4d}L +{w['presales']:3d}P -{w['cancellations']:2d}C "
                f"grL={w['gr_leads']:3d} grP={w['gr_presales']:2d} "
                f"ev={w['comm_events']:2d} "
                f"spend=${total_spend or 0:8.2f} "
                f"ig={ig_str}"
            )

            weeks_out.append(entry)

        studio_entry = {
            "name":          cfg["name"],
            "code":          code,
            "full_name":     full,
            "data_name":     cfg["data_name"],
            "targets":       cfg["targets"],
            "co_week":       co_wk,
            "go_week":       go_wk,
            "opening_date":  cfg["opening_date"].isoformat() if cfg["opening_date"] else None,
            "current_week":  cw_num,
            "weeks":         weeks_out,
        }
        output_studios.append(studio_entry)

    if dry_run:
        print("\n[DRY RUN] Not writing.")
        return

    out_path = ROOT / "nso_scorecard_data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"studios": output_studios}, f, indent=2, default=str)
    size_kb = out_path.stat().st_size // 1000
    print(f"\nOK  Written {out_path}  ({size_kb} KB)")


if __name__ == "__main__":
    main()
