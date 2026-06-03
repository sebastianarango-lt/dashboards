#!/usr/bin/env python3
"""
create_nso_config_excel.py — Generate nso_config.xlsx with all NSO studio metadata.

Run: python scripts/create_nso_config_excel.py
"""

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent.parent / "nso_config.xlsx"

# ── Brand colors ──────────────────────────────────────────────────────────────
DEEP_BLUE = "0084B5"
BLUE      = "00A3E0"
LIGHT_BG  = "EBF6FD"
MED_BG    = "D2EFFF"
WHITE     = "FFFFFF"
DARK_TEXT = "1a4d6e"

# ── Column definitions: (header_label, field_key) ────────────────────────────
SECTIONS = [
    ("STUDIO", [
        ("Name",        "name"),
        ("Code",        "code"),
        ("State",       "state"),
    ]),
    ("SCHEDULING", [
        ("Week 0 Date",      "week0"),
        ("Week 1 Start",     "week1"),
        ("Target C/O Date",  "co_date"),
        ("C/O Week #",       "co_week"),
        ("Opening Date",     "go_date"),
        ("GO Week #",        "go_week"),
        ("Tier Move Date",   "tier_move"),
    ]),
    ("TARGETS", [
        ("Total Leads",  "leads"),
        ("Presales",     "presales"),
        ("Day-1 RMR",    "rmr"),
        ("CPL Range",    "cpl"),
        ("CPA Range",    "cpa"),
        ("Conv. Rate %", "conv_rate"),
    ]),
    ("SOCIAL", [
        ("Facebook Page ID",   "fb_page"),
        ("Instagram ID",       "ig_id"),
        ("Grassroots Sheet URL", "grassroots"),
    ]),
    ("PLATFORM IDs", [
        ("Snowflake ID",   "sf_id"),
        ("Google Ads CID", "gads_cid"),
        ("GBP Location ID","gbp_id"),
    ]),
]

# ── Studio data ───────────────────────────────────────────────────────────────
STUDIOS = [
    {
        "name": "Naples - Mercato", "code": "FL-019", "state": "FL",
        "week0":    date(2026,  2,  8),
        "week1":    date(2026,  2,  9),
        "co_date":  date(2026,  7, 23), "co_week":  24,
        "go_date":  date(2026,  8, 15), "go_week":  27,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 48660,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   "986896304505624",
        "ig_id":     None,
        "grassroots": "https://docs.google.com/spreadsheets/d/1JSlGSjjRJtCqoogeAXAhli0soBW7680Gc-UInv_ZfQo/edit?usp=sharing",
        "sf_id":     5751381,
        "gads_cid":  "114-292-4076",
        "gbp_id":    "9241286551304249574",
    },
    {
        "name": "Reston", "code": "VA-001", "state": "VA",
        "week0":    date(2025, 11, 16),
        "week1":    date(2025, 11, 17),
        "co_date":  date(2026,  2,  2), "co_week":  11,
        "go_date":  None,               "go_week":  None,
        "tier_move": None,
        "leads": 929, "presales": 325, "rmr": 46425,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   "875200972337017",
        "ig_id":     "17841477453277172",
        "grassroots": "https://docs.google.com/spreadsheets/d/1G6OeztpENIK_Kr1u23OqprPuQh0UiQoNs-OtmrHetos/edit?usp=sharing",
        "sf_id":     5750130,
        "gads_cid":  "748-472-9313",
        "gbp_id":    "10767130387921211013",
    },
    {
        "name": "Herriman", "code": "UT-001", "state": "UT",
        "week0":    date(2026,  5, 17),
        "week1":    date(2026,  5, 18),
        "co_date":  date(2026,  8, 19), "co_week":  14,
        "go_date":  date(2026,  9, 12), "go_week":  17,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 48660,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   "1016504601542354",
        "ig_id":     "17841447639266583",
        "grassroots": "https://docs.google.com/spreadsheets/d/1geF_1TQXtWd5qeVCempoiP-BJRYMJHeUBZodMltUbQ0/edit?usp=sharing",
        "sf_id":     5752080,
        "gads_cid":  "385-801-4125",
        "gbp_id":    "4243744174605320602",
    },
    {
        "name": "Dr Phillips", "code": "FL-020", "state": "FL",
        "week0":    date(2026,  5, 25),
        "week1":    date(2026,  5, 26),
        "co_date":  date(2026,  8, 30), "co_week":  14,
        "go_date":  date(2026,  9, 26), "go_week":  18,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 48660,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   None,
        "ig_id":     "17841439014807825",
        "grassroots": "https://docs.google.com/spreadsheets/d/1vteVIA5_u6qodbjWRowtEr7kXsiKqS_2P8YiCnOHS8Q/edit",
        "sf_id":     5753281,
        "gads_cid":  "829-455-6178",
        "gbp_id":    "11294882594993712026",
    },
    {
        "name": "Aventura", "code": "FL-018", "state": "FL",
        "week0":    date(2026,  5, 26),
        "week1":    date(2026,  5, 27),
        "co_date":  date(2026,  8, 10), "co_week":  11,
        "go_date":  date(2026,  8, 29), "go_week":  14,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 61560,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   None,
        "ig_id":     "17841427621524852",
        "grassroots": None,
        "sf_id":     5753604,
        "gads_cid":  "959-641-5027",
        "gbp_id":    "10081543460612571163",
    },
    {
        "name": "North Miami", "code": "FL-021", "state": "FL",
        "week0":    date(2026,  5, 26),
        "week1":    date(2026,  5, 27),
        "co_date":  date(2026,  9,  1), "co_week":  14,
        "go_date":  date(2026,  9, 19), "go_week":  17,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 61560,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   None,
        "ig_id":     "17841432225694783",
        "grassroots": None,
        "sf_id":     5753608,
        "gads_cid":  None,
        "gbp_id":    "12392892117709577224",
    },
    {
        "name": "Dallas - Uptown", "code": "TX-004", "state": "TX",
        "week0":    date(2026,  5, 17),
        "week1":    date(2026,  5, 18),
        "co_date":  date(2026,  9,  7), "co_week":  16,
        "go_date":  date(2026, 10,  3), "go_week":  20,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 48660,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   None,
        "ig_id":     "17841409013036348",
        "grassroots": None,
        "sf_id":     5753491,
        "gads_cid":  "487-816-1078",
        "gbp_id":    "8299612084620782374",
    },
    {
        "name": "Old Bridge", "code": "NJ-004", "state": "NJ",
        "week0":    date(2026,  7, 27),
        "week1":    date(2026,  7, 28),
        "co_date":  date(2026, 11,  1), "co_week":  14,
        "go_date":  date(2026, 11, 21), "go_week":  17,
        "tier_move": None,
        "leads": 1257, "presales": 440, "rmr": 48660,
        "cpl": "28 - 41", "cpa": "80 - 116", "conv_rate": 35,
        "fb_page":   None,
        "ig_id":     "17841439161726674",
        "grassroots": None,
        "sf_id":     5753073,
        "gads_cid":  None,
        "gbp_id":    "9821491381105118276",
    },
]

# ── Column widths (one per field in order) ────────────────────────────────────
COL_WIDTHS = [
    26, 10, 7,                    # Studio
    14, 14, 16, 10, 14, 10, 16,   # Scheduling
    11, 11, 13, 11, 11, 10,       # Targets (added conv_rate)
    22, 22, 34,                   # Social
    14, 14, 24,                   # Platform IDs
]

DATE_FMT = "MM/DD/YYYY"
THIN = Side(style="thin", color="B0C8D8")


def border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NSO Config"

    # ── Flatten sections ──────────────────────────────────────────────────────
    flat_cols = []
    section_spans = []
    for sec_name, fields in SECTIONS:
        start = len(flat_cols) + 1
        for label, key in fields:
            flat_cols.append((label, key))
        section_spans.append((sec_name, start, len(flat_cols)))

    # Row 1 — section headers (merged)
    for sec_name, c1, c2 in section_spans:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=sec_name)
        cell.font = Font(name="Calibri", bold=True, italic=True, color=WHITE, size=12)
        cell.fill = PatternFill("solid", fgColor=DEEP_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border()

    # Row 2 — column headers
    for ci, (label, _) in enumerate(flat_cols, 1):
        cell = ws.cell(row=2, column=ci, value=label)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border()

    # Data rows
    row_bgs = [LIGHT_BG, MED_BG]
    for si, studio in enumerate(STUDIOS):
        row = si + 3
        bg = row_bgs[si % 2]
        for ci, (label, key) in enumerate(flat_cols, 1):
            val = studio.get(key)
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = border()
            halign = "center"
            bold = False
            if ci == 1:  # Name
                halign = "left"
                bold = True
            elif key in ("grassroots",):
                halign = "left"
            cell.font = Font(name="Calibri", color=DARK_TEXT, size=11, bold=bold)
            cell.alignment = Alignment(horizontal=halign, vertical="center")
            if isinstance(val, date):
                cell.number_format = DATE_FMT

    # ── Dimensions ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    for r in range(3, 3 + len(STUDIOS)):
        ws.row_dimensions[r].height = 26

    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    wb.save(OUT)
    print(f"Created {OUT}")


if __name__ == "__main__":
    build()
